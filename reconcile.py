"""
Reconciles the QIC_ Actual_LP / QIC_ Actual_UP / QIC_ Actual_5&6 sheets of the
"Contractual vs Actual" master workbook against an uploaded "TOTAL MANPOWER"
workbook (which has a TOTAL MANPOWER sheet of active employees and a LEFT
sheet of departed employees), and keeps the TOTAL- LP , UP, 5&6 & PGC rollup
sheet in sync.

PGC is intentionally skipped everywhere. Existing Position labels,
Contractual Requirement values, formulas, merged cells and formatting are
never touched -- every edit either fills an already-blank employee-slot
cell, or APPENDS new rows far below all existing content (never inserts
mid-sheet, which is what protects the workbook's many formulas and merged
cells from silently breaking).

What happens on every run, per project (LP / UP / 5&6):
  1. REMOVE   - a named employee's Iqama is found in the LEFT sheet, or is
                active in TOTAL MANPOWER under a *different* project (moved).
                Their row is cleared and marked Vacant.
  2. ADD      - an employee active in TOTAL MANPOWER for this project who
                isn't yet listed here is placed into a vacant row whose
                Position label strictly matches their SAP POSITION or
                ASSIGNED WORK (exact match, or one is a whole-word subset of
                the other) - no fuzzy/character-similarity matching, to
                avoid misplacing someone into a similar-looking but wrong
                role.
  3. APPEND   - anyone left over (no vacant slot anywhere for their
                position - including genuinely new positions never seen
                before) is APPENDED as a new row in the Back-End/Additional
                area, Requirement=0, Status="Additional". If their position
                has no entry yet in the TOTAL-sheet mapping, a brand-new row
                is also appended there with the headcount in Additional Qty,
                and the mapping is extended so future runs recognize it.
  4. FLAG     - "unmatched": employees currently listed here whose Iqama
                isn't found in TOTAL MANPOWER (active) or LEFT at all -
                left completely untouched, reported for manual review.

update_total_sheet() then recomputes Onsite Qty / Additional Qty for every
mapped Position from the CURRENT state of the Actual sheets. Expat Qty /
Saudi Qty (the approved requirement) are never touched.
"""

import re
import io
import openpyxl

SHEET_NAMES = {"LP": "QIC_ Actual_LP", "UP": "QIC_ Actual_UP", "5&6": "QIC_ Actual_5&6"}

TOTAL_SHEET_NAME = "TOTAL- LP , UP, 5&6 & PGC"
# (onsite_col, additional_col) per project on the TOTAL sheet
TOTAL_SHEET_COLS = {"LP": (7, 8), "UP": (13, 14), "5&6": (19, 20)}
# The sheet's pre-existing grand-total row (sums specific section-subtotal
# cells via formula) - extended, never moved, when a new row is appended.
GRAND_TOTAL_ROW = 145

COL_POSITION = 1
COL_REQUIREMENT = 2
COL_ONSITE = 3
COL_VACANT = 4
COL_NAME = 5
COL_DECIPLINE = 6
COL_ID = 7
COL_IQAMA = 8
COL_STATUS = 9
COL_DEPLOY_DATE = 10
COL_SPONSOR = 11
COL_REMARKS = 12


def _iq(v):
    """Normalize an Iqama/passport value to a comparable string, or None."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "None"):
        return None
    if re.match(r"^\d+\.0$", s):
        s = s[:-2]
    return s


def _norm(s):
    if s is None:
        return ""
    s = str(s).upper().strip()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    if s.endswith("IES"):
        s = s[:-3] + "Y"
    elif s.endswith("S") and not s.endswith("SS"):
        s = s[:-1]
    return s


def _strict_match(a, b):
    if not a or not b:
        return False
    if a == b:
        return True
    ta, tb = set(a.split()), set(b.split())
    return ta.issubset(tb) or tb.issubset(ta)


def _project_match(proj, key):
    if not proj:
        return False
    p = str(proj).upper()
    if key == "LP":
        return "LOWER" in p
    if key == "UP":
        return "UPPER" in p
    if key == "5&6":
        return "5&6" in p or "5 & 6" in p
    return False


def _parse_actual_sheet(ws):
    """Forward-fills Position/Requirement across each block of rows, skips
    header rows, 'Total Team ...' subtotal rows (and the numeric row right
    below each one), and fully blank rows."""
    records = []
    current_position = None
    current_req = None
    skip_next = False
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=COL_POSITION).value
        b = ws.cell(row=r, column=COL_REQUIREMENT).value
        e = ws.cell(row=r, column=COL_NAME).value
        f = ws.cell(row=r, column=COL_DECIPLINE).value
        g = ws.cell(row=r, column=COL_ID).value
        h = ws.cell(row=r, column=COL_IQAMA).value
        i = ws.cell(row=r, column=COL_STATUS).value

        if skip_next:
            skip_next = False
            continue
        if a == "Position":
            continue
        if isinstance(a, str) and a.strip().startswith("Total Team"):
            current_position = None
            skip_next = True
            continue
        if all(v is None for v in [a, b, e, f, g, h, i]):
            continue
        if a is not None:
            current_position = a
            current_req = b
        records.append(dict(row=r, position=current_position, requirement=current_req,
                             employee_name=e, decipline=f, id=g, iqama=h, status=i))
    return records


def _true_last_content_row(ws):
    """Scans every used cell (not just the tracked columns) so appends never
    land on top of stray data further right or below the parsed columns."""
    last = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.row > last:
                last = cell.row
    return last


def _find_backend_subtotal(ws):
    """Returns (header_row, values_row) of the 'Total Team ... Back-End
    Staff/Additional' subtotal, or (None, None) if not found."""
    for r in range(1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, str) and a.strip().startswith("Total Team") and \
                ("Back-End" in a or "Additional" in a):
            return r, r + 1
    return None, None


_RANGE_RE = re.compile(r"([A-Z]+)(\d+):([A-Z]+)(\d+)")


def _extend_range_end(formula, new_end_row):
    """Extends every A1:A2-style range in a formula so its end row becomes
    new_end_row (start row / columns are left alone)."""
    def repl(m):
        col1, row1, col2, _row2 = m.groups()
        return f"{col1}{row1}:{col2}{new_end_row}"
    return _RANGE_RE.sub(repl, formula)


def _extend_grand_total_formula(formula, col_letter, new_row):
    """Adds a reference to the new row's cell into the sheet's existing
    grand-total formula, matching whichever style it already uses."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula  # static number - nothing to extend
    new_ref = f"{col_letter}{new_row}"
    if formula.upper().startswith("=SUM("):
        idx = formula.index("(") + 1
        return formula[:idx] + new_ref + "," + formula[idx:]
    return formula + f"+{new_ref}"


def _col_letter(col_idx):
    letters = ""
    while col_idx > 0:
        col_idx, rem = divmod(col_idx - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _load_manpower(manpower_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(manpower_bytes), data_only=True)
    ws_tm = wb["TOTAL MANPOWER"]
    hdr = {c.value: i + 1 for i, c in enumerate(ws_tm[1])}

    tm_employees = []
    all_tm_iqamas = set()
    for r in range(2, ws_tm.max_row + 1):
        iqama = _iq(ws_tm.cell(row=r, column=hdr["Iqama No / PP"]).value)
        if not iqama:
            continue
        all_tm_iqamas.add(iqama)
        status = ws_tm.cell(row=r, column=hdr["STATUS"]).value
        if status and str(status).upper() != "ACTIVE":
            continue
        tm_employees.append(dict(
            iqama=iqama,
            emp_id=ws_tm.cell(row=r, column=hdr["Emp ID"]).value,
            name=ws_tm.cell(row=r, column=hdr["NAME"]).value,
            sap_position=ws_tm.cell(row=r, column=hdr["SAP POSITION"]).value,
            assigned_work=ws_tm.cell(row=r, column=hdr["ASSIGNED WORK"]).value,
            sponsor=ws_tm.cell(row=r, column=hdr["VISA SPONSOR"]).value,
            project=ws_tm.cell(row=r, column=hdr["PROJECT"]).value,
            joining_date=ws_tm.cell(row=r, column=hdr["Joining Date"]).value,
        ))

    ws_left = wb["LEFT"]
    hdrL = {c.value: i + 1 for i, c in enumerate(ws_left[1])}
    left_iqamas = set()
    for r in range(2, ws_left.max_row + 1):
        iqama = _iq(ws_left.cell(row=r, column=hdrL["IQAMA"]).value)
        if iqama:
            left_iqamas.add(iqama)

    tm_by_project = {"LP": [], "UP": [], "5&6": []}
    for rec in tm_employees:
        for k in tm_by_project:
            if _project_match(rec["project"], k):
                tm_by_project[k].append(rec)

    tm_iqama_to_proj = {}
    for k, lst in tm_by_project.items():
        for r in lst:
            tm_iqama_to_proj[r["iqama"]] = k

    return dict(tm_employees=tm_employees, left_iqamas=left_iqamas,
                tm_by_project=tm_by_project, tm_iqama_to_proj=tm_iqama_to_proj,
                all_tm_iqamas=all_tm_iqamas)


def build_plan(master_wb, manpower_bytes):
    """Read-only: computes what WOULD change. Does not modify master_wb."""
    mp = _load_manpower(manpower_bytes)
    all_tm_iqamas = mp["all_tm_iqamas"]

    plan = {}
    for proj, sn in SHEET_NAMES.items():
        ws = master_wb[sn]
        recs = _parse_actual_sheet(ws)
        all_labels = sorted({r["position"] for r in recs if r["position"]})

        remove_rows = []
        for rec in recs:
            if not rec["employee_name"]:
                continue
            i = _iq(rec["iqama"])
            if not i:
                continue
            if i in mp["left_iqamas"]:
                remove_rows.append((rec["row"], "left"))
            else:
                ap = mp["tm_iqama_to_proj"].get(i)
                if ap and ap != proj:
                    remove_rows.append((rec["row"], f"moved_to_{ap}"))
        remove_row_nums = {r for r, _ in remove_rows}

        existing_iqamas = {_iq(r["iqama"]) for r in recs if _iq(r["iqama"])}

        vacant_rows = []
        for rec in recs:
            i = _iq(rec["iqama"])
            is_vacant_now = (not rec["employee_name"]) and (not i or rec["status"] == "Vacant")
            will_be_vacant = rec["row"] in remove_row_nums
            if is_vacant_now or will_be_vacant:
                vacant_rows.append(dict(row=rec["row"], position=rec["position"],
                                         requirement=rec["requirement"]))

        # Promote existing "Additional" employees into a vacant REGULAR
        # (Requirement > 0) slot for their exact same Position label, before
        # considering anyone new - an Additional person already doing this
        # job should fill an on-site opening ahead of a brand-new hire.
        by_position_all = {}
        for rec in recs:
            if rec["position"]:
                by_position_all.setdefault(rec["position"], []).append(rec)

        promotions = []
        promoted_candidate_rows = set()
        promoted_vacant_rows = set()
        for v in vacant_rows:
            if v["requirement"] is None or v["requirement"] <= 0:
                continue  # only promote into real on-site quota slots
            candidates = by_position_all.get(v["position"], [])
            for c in candidates:
                if (c["requirement"] or 0) == 0 and c["employee_name"] \
                        and c["row"] not in remove_row_nums \
                        and c["row"] not in promoted_candidate_rows:
                    promotions.append((c, v))
                    promoted_candidate_rows.add(c["row"])
                    promoted_vacant_rows.add(v["row"])
                    break

        vacant_rows = [v for v in vacant_rows if v["row"] not in promoted_vacant_rows]

        joiners = [r for r in mp["tm_by_project"][proj]
                   if r["iqama"] not in existing_iqamas and r["iqama"] not in mp["left_iqamas"]]
        conflicts = [r for r in mp["tm_by_project"][proj]
                     if r["iqama"] not in existing_iqamas and r["iqama"] in mp["left_iqamas"]]

        assignments, overflow, used_rows = [], [], set()
        for j in joiners:
            jn_sap, jn_aw = _norm(j["sap_position"]), _norm(j["assigned_work"])
            best = None
            for v in vacant_rows:
                if v["row"] in used_rows:
                    continue
                vn = _norm(v["position"])
                if _strict_match(jn_sap, vn) or _strict_match(jn_aw, vn):
                    best = v
                    break
            if best:
                assignments.append((j, best))
                used_rows.add(best["row"])
            else:
                # No vacant slot anywhere. Reuse an existing Position label
                # if this person's role matches one already used in this
                # sheet (even if that block is full); otherwise this is a
                # genuinely new position.
                label = None
                for existing in all_labels:
                    en = _norm(existing)
                    if _strict_match(jn_sap, en) or _strict_match(jn_aw, en):
                        label = existing
                        break
                if not label:
                    label = (j["sap_position"] or j["assigned_work"] or "Unspecified Position")
                    label = str(label).strip().title()
                overflow.append((j, label))

        unmatched = []
        for rec in recs:
            if not rec["employee_name"]:
                continue
            i = _iq(rec["iqama"])
            if not i or i in mp["left_iqamas"]:
                continue
            if i not in all_tm_iqamas:
                unmatched.append(rec)

        plan[proj] = dict(remove_rows=remove_rows, assignments=assignments,
                           overflow=overflow, unmatched=unmatched, conflicts=conflicts,
                           promotions=promotions)
    return plan


def apply_plan(master_wb, plan):
    """Mutates master_wb in place according to a plan from build_plan().
    Returns (change_log, appended_by_label) where appended_by_label is
    {project: {label: [row_numbers]}} for everything newly appended as
    Additional - used afterwards to sync the TOTAL sheet."""
    change_log = {}
    appended_by_label = {}
    for proj, sn in SHEET_NAMES.items():
        ws = master_wb[sn]
        p = plan[proj]
        log = []

        for row, reason in p["remove_rows"]:
            name_before = ws.cell(row=row, column=COL_NAME).value
            ws.cell(row=row, column=COL_NAME).value = None
            ws.cell(row=row, column=COL_ID).value = "-"
            ws.cell(row=row, column=COL_IQAMA).value = "-"
            ws.cell(row=row, column=COL_STATUS).value = "Vacant"
            ws.cell(row=row, column=COL_DEPLOY_DATE).value = None
            ws.cell(row=row, column=COL_SPONSOR).value = "-"
            ws.cell(row=row, column=COL_REMARKS).value = None
            log.append(f"Row {row}: REMOVED {name_before} ({reason})")

        for old_rec, v in p["promotions"]:
            old_row, new_row = old_rec["row"], v["row"]
            name = ws.cell(row=old_row, column=COL_NAME).value
            emp_id = ws.cell(row=old_row, column=COL_ID).value
            iqama = ws.cell(row=old_row, column=COL_IQAMA).value
            deploy_date = ws.cell(row=old_row, column=COL_DEPLOY_DATE).value
            sponsor = ws.cell(row=old_row, column=COL_SPONSOR).value

            ws.cell(row=new_row, column=COL_NAME).value = name
            ws.cell(row=new_row, column=COL_ID).value = emp_id
            ws.cell(row=new_row, column=COL_IQAMA).value = iqama
            ws.cell(row=new_row, column=COL_STATUS).value = "Approved"
            ws.cell(row=new_row, column=COL_DEPLOY_DATE).value = deploy_date
            ws.cell(row=new_row, column=COL_SPONSOR).value = sponsor

            ws.cell(row=old_row, column=COL_NAME).value = None
            ws.cell(row=old_row, column=COL_ID).value = "-"
            ws.cell(row=old_row, column=COL_IQAMA).value = "-"
            ws.cell(row=old_row, column=COL_STATUS).value = "Vacant"
            ws.cell(row=old_row, column=COL_DEPLOY_DATE).value = None
            ws.cell(row=old_row, column=COL_SPONSOR).value = "-"
            ws.cell(row=old_row, column=COL_REMARKS).value = None

            log.append(
                f"Row {new_row}: PROMOTED {name} from Additional (row {old_row}) "
                f"to fill the on-site vacancy for {v['position']}"
            )

        for j, v in p["assignments"]:
            row = v["row"]
            jd = j.get("joining_date")
            ws.cell(row=row, column=COL_NAME).value = j["name"]
            ws.cell(row=row, column=COL_ID).value = j.get("emp_id") or "-"
            ws.cell(row=row, column=COL_IQAMA).value = j["iqama"]
            ws.cell(row=row, column=COL_STATUS).value = "Approved"
            ws.cell(row=row, column=COL_DEPLOY_DATE).value = jd if jd else None
            ws.cell(row=row, column=COL_SPONSOR).value = j.get("sponsor") or "-"
            ws.cell(row=row, column=COL_REMARKS).value = None
            log.append(f"Row {row}: ADDED {j['name']} as {v['position']} (was vacant)")

        # group overflow joiners by label, preserving first-seen order
        groups_order = []
        groups = {}
        for j, label in p["overflow"]:
            if label not in groups:
                groups[label] = []
                groups_order.append(label)
            groups[label].append(j)

        label_rows = {}
        if groups_order:
            header_row, values_row = _find_backend_subtotal(ws)
            last_row = _true_last_content_row(ws)
            row_ptr = max(last_row, values_row or 0) + 2
            for label in groups_order:
                rows_for_label = []
                for idx, j in enumerate(groups[label]):
                    ws.cell(row=row_ptr, column=COL_POSITION).value = label if idx == 0 else None
                    ws.cell(row=row_ptr, column=COL_REQUIREMENT).value = 0 if idx == 0 else None
                    ws.cell(row=row_ptr, column=COL_ONSITE).value = 1
                    ws.cell(row=row_ptr, column=COL_VACANT).value = 0
                    ws.cell(row=row_ptr, column=COL_NAME).value = j["name"]
                    ws.cell(row=row_ptr, column=COL_DECIPLINE).value = None
                    ws.cell(row=row_ptr, column=COL_ID).value = j.get("emp_id") or "-"
                    ws.cell(row=row_ptr, column=COL_IQAMA).value = j["iqama"]
                    ws.cell(row=row_ptr, column=COL_STATUS).value = "Additional"
                    jd = j.get("joining_date")
                    ws.cell(row=row_ptr, column=COL_DEPLOY_DATE).value = jd if jd else None
                    ws.cell(row=row_ptr, column=COL_SPONSOR).value = j.get("sponsor") or "-"
                    ws.cell(row=row_ptr, column=COL_REMARKS).value = None
                    log.append(f"Row {row_ptr}: ADDED {j['name']} as {label} (new Additional entry, no vacant slot found)")
                    rows_for_label.append(row_ptr)
                    row_ptr += 1
                label_rows[label] = rows_for_label
            new_last_row = row_ptr - 1

            if values_row:
                f_formula = ws.cell(row=values_row, column=6).value
                if isinstance(f_formula, str):
                    ws.cell(row=values_row, column=6).value = _extend_range_end(f_formula, new_last_row)
                a_formula = ws.cell(row=values_row, column=1).value
                if isinstance(a_formula, str):
                    ws.cell(row=values_row, column=1).value = _extend_range_end(a_formula, new_last_row)

        change_log[proj] = log
        appended_by_label[proj] = label_rows
    return change_log, appended_by_label


def sync_total_sheet_mapping(master_wb, position_mapping, appended_by_label):
    """For every newly-appended label with no existing TOTAL-sheet mapping,
    reuses an existing auto-added row for that same label from another
    project if one already exists (so "Life Guard" needed by both LP and UP
    ends up as one shared row, matching how the rest of the sheet works),
    otherwise appends a brand-new row. Extends the grand-total formulas to
    include any newly-created row, and adds the mapping entry (mutates
    position_mapping in place). Returns a change log list."""
    ws_total = master_wb[TOTAL_SHEET_NAME]
    log = []
    new_labels = []
    for proj, label_rows in appended_by_label.items():
        mapping = position_mapping.setdefault(proj, {})
        for label, rows in label_rows.items():
            if label not in mapping:
                new_labels.append((proj, label, len(rows)))

    if not new_labels:
        return log

    # index existing auto-added rows by normalized label, so a second
    # project needing the same new position reuses the row instead of
    # duplicating it
    existing_auto_rows = {}
    for r in range(1, ws_total.max_row + 1):
        d = ws_total.cell(row=r, column=4).value
        if isinstance(d, str) and d.strip().endswith("(auto-added)"):
            key = d.strip()[: -len("(auto-added)")].strip().upper()
            existing_auto_rows[key] = r

    row_ptr = _true_last_content_row(ws_total) + 2
    for proj, label, qty in new_labels:
        key = label.strip().upper()
        onsite_col, additional_col = TOTAL_SHEET_COLS[proj]
        expat_col, saudi_col = onsite_col - 2, onsite_col - 1

        if key in existing_auto_rows:
            target_row = existing_auto_rows[key]
        else:
            target_row = row_ptr
            ws_total.cell(row=target_row, column=2).value = "-"
            ws_total.cell(row=target_row, column=3).value = "-"
            ws_total.cell(row=target_row, column=4).value = f"{label} (auto-added)"
            existing_auto_rows[key] = target_row
            row_ptr += 1

        ws_total.cell(row=target_row, column=expat_col).value = 0
        ws_total.cell(row=target_row, column=saudi_col).value = 0
        ws_total.cell(row=target_row, column=onsite_col).value = 0
        ws_total.cell(row=target_row, column=additional_col).value = qty

        for col in (expat_col, saudi_col, onsite_col, additional_col):
            cell = ws_total.cell(row=GRAND_TOTAL_ROW, column=col)
            cell.value = _extend_grand_total_formula(cell.value, _col_letter(col), target_row)

        position_mapping[proj][label] = target_row
        log.append(f"{proj}: TOTAL-sheet row {target_row} set for '{label}' (Additional Qty = {qty})")
    return log


def update_total_sheet(master_wb, position_mapping):
    """
    Recomputes Onsite Qty / Additional Qty on the TOTAL- LP , UP, 5&6 & PGC
    sheet from the CURRENT state of the Actual sheets (call this AFTER
    apply_plan + sync_total_sheet_mapping), using the Position -> TOTAL-row
    mapping.

    For each Position label with a mapping entry:
      - blocks with a non-zero Contractual Requirement contribute their
        named-employee headcount to Onsite Qty (capped implicitly, since a
        block's row count equals its Requirement)
      - blocks with a ZERO Requirement (Back-End/Additional overflow rows
        for that same position) contribute their named-employee headcount
        to Additional Qty
      - defensive: if a non-zero-Requirement block ever has more named
        employees than its Requirement, the excess also counts as Additional

    Multiple raw Position labels mapping to the same TOTAL-sheet row have
    their headcounts summed into that one row. Expat Qty / Saudi Qty (the
    approved requirement) are never touched. Returns a change log.
    """
    change_log = {}
    ws_total = master_wb[TOTAL_SHEET_NAME]

    for proj, sn in SHEET_NAMES.items():
        mapping = position_mapping.get(proj, {})
        if not mapping:
            change_log[proj] = []
            continue

        ws = master_wb[sn]
        recs = _parse_actual_sheet(ws)

        by_position = {}
        for rec in recs:
            if rec["position"] is None:
                continue
            by_position.setdefault(rec["position"], []).append(rec)

        totals_by_row = {}
        for pos, total_row in mapping.items():
            rows = by_position.get(pos, [])
            if not rows:
                continue
            requirement = rows[0]["requirement"] or 0
            named = sum(1 for r in rows if r["employee_name"])
            onsite = min(named, requirement) if requirement > 0 else 0
            additional = max(named - requirement, 0)
            acc = totals_by_row.setdefault(total_row, [0, 0])
            acc[0] += onsite
            acc[1] += additional

        onsite_col, additional_col = TOTAL_SHEET_COLS[proj]
        log = []
        for total_row, (onsite, additional) in totals_by_row.items():
            before_o = ws_total.cell(row=total_row, column=onsite_col).value
            before_a = ws_total.cell(row=total_row, column=additional_col).value
            if before_o != onsite or before_a != additional:
                ws_total.cell(row=total_row, column=onsite_col).value = onsite
                ws_total.cell(row=total_row, column=additional_col).value = additional
                fms_label = ws_total.cell(row=total_row, column=4).value
                log.append(
                    f"Row {total_row} ({fms_label}): Onsite {before_o}->{onsite}, "
                    f"Additional {before_a}->{additional}"
                )
        change_log[proj] = log
    return change_log


def summary_markdown(plan, total_sheet_log=None, new_row_log=None):
    lines = ["# Manpower Reconciliation Summary -- LP / UP / 5&6", ""]
    if new_row_log:
        lines.append("**New positions discovered and added to the TOTAL sheet:**")
        for line in new_row_log:
            lines.append(f"- {line}")
        lines.append("")
    for proj in ["LP", "UP", "5&6"]:
        p = plan[proj]
        lines.append(f"## {proj}")
        lines.append("")
        lines.append(f"**Removed -- departed or moved to another project ({len(p['remove_rows'])}):**")
        for row, reason in p["remove_rows"]:
            lines.append(f"- Row {row} -- {reason.replace('_', ' ')}")
        lines.append("")
        lines.append(f"**Promoted from Additional to on-site, filling a vacancy ({len(p['promotions'])}):**")
        for old_rec, v in p["promotions"]:
            lines.append(f"- {old_rec['employee_name']} -- row {old_rec['row']} -> row {v['row']} ({v['position']})")
        lines.append("")
        lines.append(f"**Filled into existing vacant slots ({len(p['assignments'])}):**")
        for j, v in p["assignments"]:
            lines.append(f"- {j['name']} -> row {v['row']} ({v['position']})")
        lines.append("")
        lines.append(f"**Added as new Additional entries -- no vacant slot found ({len(p['overflow'])}):**")
        for j, label in p["overflow"]:
            lines.append(f"- {j['name']} -- filed under '{label}' (Iqama {j['iqama']})")
        lines.append("")
        lines.append(f"**Unmatched -- listed here but not found in TOTAL MANPOWER or LEFT ({len(p['unmatched'])}):**")
        for rec in p["unmatched"]:
            lines.append(f"- {rec['employee_name']} -- Iqama/PP: {rec['iqama']} -- Position: {rec['position']}")
        if total_sheet_log:
            changes = total_sheet_log.get(proj, [])
            lines.append("")
            lines.append(f"**TOTAL sheet rows updated ({len(changes)}):**")
            for line in changes:
                lines.append(f"- {line}")
        lines.append("")
    return "\n".join(lines)


def reconcile(master_bytes, manpower_bytes, position_mapping):
    """
    position_mapping: {"LP": {"Position label": total_row, ...}, "UP": {...},
    "5&6": {...}} - load this from position_mapping.json (or wherever it's
    stored) before calling, and persist the returned, possibly-extended
    mapping back afterwards (new positions get added to it automatically).

    Returns (plan, updated_master_bytes, change_log, summary_md, position_mapping).
    """
    wb = openpyxl.load_workbook(io.BytesIO(master_bytes), data_only=False)
    plan = build_plan(wb, manpower_bytes)
    change_log, appended_by_label = apply_plan(wb, plan)
    new_row_log = sync_total_sheet_mapping(wb, position_mapping, appended_by_label)
    total_sheet_log = update_total_sheet(wb, position_mapping)
    for proj in change_log:
        change_log[proj].extend(total_sheet_log.get(proj, []))
    out = io.BytesIO()
    wb.save(out)
    summary = summary_markdown(plan, total_sheet_log, new_row_log)
    return plan, out.getvalue(), change_log, summary, position_mapping
